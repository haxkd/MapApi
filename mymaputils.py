from geopy.geocoders import Nominatim
import csv
import openpyxl
import os
ALLOWED_EXTENSIONS = set(['xlsx','xlsm','xltx','xltm','csv'])


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS



def getmapdata(file_path,filename):
    locator = Nominatim(user_agent='myGeocoder')
    all_rows_data = []
    list_data = []

    wb_obj = openpyxl.load_workbook(file_path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    if m_row<1:
        return False

    for i in range(2, m_row + 1):
        cell_obj_long = sheet_obj.cell(row = i, column = 1)
        cell_obj_lat = sheet_obj.cell(row = i, column = 2)
        list_data.append((cell_obj_long.value,cell_obj_lat.value))
    #print(list_data)
    
    for row in list_data:
        location = locator.reverse(row)
        if location ==None:
            all_rows_data.append([row[0],row[1],'address not Found'])
            continue
        all_rows_data.append([row[0],row[1],location.address])
    print(all_rows_data)
    if len(all_rows_data)>0:
        with open(os.path.join('static/uploads',filename+'.csv'), 'w', newline='') as new_data_file:
            csvwriter = csv.writer(new_data_file)
            csvwriter.writerow(['longitude','lattitude','address'])
            # writing the data rows
            csvwriter.writerows(all_rows_data)
            print(csvwriter,csvwriter,'hellllloooooooooo')
    return True


def getmapdatacsv(file_path):
    locator = Nominatim(user_agent='myGeocoder')
    all_rows_data = []
    # file_path = 'uploads/coord.csv'
    with open(file_path) as cordfile:
        csvreader = csv.reader(cordfile)
        fields = next(csvreader)
        for row in csvreader:
            c = (float(row[0].strip()),float(row[1].strip()))
            location = locator.reverse(c)
            all_rows_data.append([row[0].strip(),row[1].strip(),location.address])

    if len(all_rows_data)>0:
        with open(file_path, 'w', newline='') as new_data_file:
            csvwriter = csv.writer(new_data_file)
            csvwriter.writerow(['longitude','lattitude','address'])
            # writing the data rows
            csvwriter.writerows(all_rows_data)
    return True